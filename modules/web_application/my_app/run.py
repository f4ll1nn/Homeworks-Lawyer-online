""" Running web application """
from flask import Flask, render_template, request
import openpyxl

APP = Flask(__name__)


@APP.route("/")
def index():
    """ Running start.html"""
    return render_template("start.html")


@APP.route("/", methods=["POST"])
def pars_urls():
    """ Saving result from form """
    situation = request.form["message"].strip()
    category = request.form["dept"]
    name = request.form["name"]
    email = request.form["email"]
    workbook = openpyxl.load_workbook("result.xlsx")
    worksheet = workbook.active
    worksheet.cell(column=1, row=worksheet.max_row+1, value=situation)
    worksheet.cell(column=2, row=worksheet.max_row, value=category)
    worksheet.cell(column=3, row=worksheet.max_row, value=name)
    worksheet.cell(column=4, row=worksheet.max_row, value=email)
    workbook.save("result.xlsx")
    return render_template("result.html")


if __name__ == "__main__":
    APP.run(debug=True)
