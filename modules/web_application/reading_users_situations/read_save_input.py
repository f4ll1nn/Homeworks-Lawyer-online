""" Realization usersituation upworking """
import os

import openpyxl
import xlsxwriter

from kku.trans.mtranslate.mtranslate import translate


class UsersSituation:
    """ Class of users situations """

    def __init__(self, data=None, category=None, result=None):
        """ Creates users situations """
        self.data = data
        self.category = category
        self.result = result
        self.translated_situation = None

    def read_users_situation(self, filename: str) -> str:
        """ Reads users_situation from exel file. Returns users situation """
        path = os.path.normpath(os.getcwd() + os.sep + os.pardir)
        os.chdir(path)
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        self.data = worksheet.cell(column=1, row=worksheet.max_row).value
        self.category = worksheet.cell(column=2, row=worksheet.max_row).value
        return self.data

    def history_user_write(self) -> None:
        """ Creates file for current user history with his situation """
        writer = xlsxwriter.Workbook("users_history.xlsx")
        worksheet = writer.add_worksheet()
        row = 0
        worksheet.write(row, 0, self.data)
        worksheet.write(row, 1, self.category)
        writer.close()

    def translate_situation(self) -> None:
        """ Translates situation into english """
        translated = translate(self.data, "en")
        self.translated_situation = translated
        return self.translated_situation

    def choicing_relevant_information_from_situation(self):
        """ Chose relevant information from situation discription """
        pass

    def find_persons(self):
        """ Finds main persons in situaion """
        pass

    def find_articles(self):
        """ Find articles ro situations """
        pass

    def saves_result(self, filename1: str, filename2: str) -> None:
        """ Writes result to users_history and to result """
        workbook = openpyxl.load_workbook(filename1)
        worksheet = workbook.active
        worksheet.cell(column=worksheet.max_column + 1, row=worksheet.max_row,
                       value=self.result)
        workbook.save(filename1)
        workbook = openpyxl.load_workbook(filename2)
        worksheet = workbook.active
        worksheet.cell(column=worksheet.max_column + 1, row=worksheet.max_row,
                       value=self.result)
        workbook.save(filename2)


if __name__ == "__main__":
    USER_SITUATION = UsersSituation()
    USER_SITUATION.read_users_situation("result.xlsx")
    USER_SITUATION.history_user_write()
    print(USER_SITUATION.translate_situation())
